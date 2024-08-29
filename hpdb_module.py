#----------------------
# Import Library & Workbook
#----------------------
import time
from colorama import Fore, Style
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
from datetime import datetime
import os
from kmz_module import getAllHP, getAllFAT
from ForceSys import formatted_print

hpdb_col = [
        'ACQUISITION_CLASS',
        'ACQUISITION_TIER',
        'BUILDING_TYPE',
        'OWNERSHIP',
        'VENDOR_NAME',
        'ZIP_CODE',
        'REGION',
        'CITY',
        'CITY_CODE',
        'DISTRICT',
        'SUB_DISTRICT',
        'FAT_CODE',
        'FAT_LONGITUDE',
        'FAT_LATITUDE',
        'BUILDING_LATITUDE',
        'BUILDING_LONGITUDE',
        'HOMEPASS_ID',
        'MOBILE_REGION',
        'MOBILE_CLUSTER',
        'CITY_GROUP'
    ]

log_location = '   HPDB'
text_color = Fore.LIGHTMAGENTA_EX

def get_excel_sheet_names(filename):
    try:
        # Load the Excel workbook
        workbook = load_workbook(filename)

        # Get all sheet names
        sheet_names = workbook.sheetnames

        return sheet_names

    except FileNotFoundError:
        formatted_print(f'File not found', "   HPDB", Fore.LIGHTMAGENTA_EX)
        return None
        
def find_matching_strings(pattern, string_list):
    matching_strings = []
    for string in string_list:
        if re.search(pattern, string, re.IGNORECASE):
            matching_strings.append(string)
    if len(matching_strings) == 0:
        for string in string_list:
            if "HPDB" in string:
                matching_strings.append(string)
    return matching_strings

def date_converter(value):
    return value.strftime('%d/%m/%Y') if isinstance(value, pd.Timestamp) else value

# Process each file
def hpdbCheck(raw_file_path, kmz_file_path, cluster, checking_date, checking_time, kmz_col, log_col):

    start_time = time.time()
    file = os.path.basename(raw_file_path)
    formatted_print(f'Checking HPDB File: {file}', log_location, text_color)
    try:
        homepass_long_lat = getAllHP(kmz_file_path)
        has_hp = True
    except:
        has_hp = False
        formatted_print(f'No Homepass folder found', log_location, text_color)

    try:
        fat_long_lat = getAllFAT(kmz_file_path)
        has_fat = True
    except:
        has_fat = False
        formatted_print(f'No FAT folder found', log_location, text_color)

    # Define the directory paths
    output_dir = 'Output'
    summary_dir = f"Summary"
    summary_file_path = os.path.join(summary_dir, f"Summary_{checking_date}.csv")
    summary_file_path_cluster = os.path.join(summary_dir, f"{cluster}.csv")

    # Construct the full file paths
    output_file_path = os.path.join(output_dir, f"output_{file}")

    # Load the "kodepos.xlsx" Excel file into a pandas DataFrame
    kodepos_df = pd.read_excel('Reference/ZIP_Code.xlsx')

    # Load the "Mobile_Region_Cluster.xlsx" Excel file into a pandas DataFrame
    mobile_df = pd.read_excel('Reference/Mobile_Region_Cluster.xlsx')
    
    sheet_names = get_excel_sheet_names(raw_file_path)

    if len(sheet_names) == 1:
        hpdb_sheet = sheet_names
    
    else:
        # Find matching strings
        pattern = r"^HPDB"
        hpdb_sheet = find_matching_strings(pattern, sheet_names)
    
    # Load the file into a pandas DataFrame
    hpdb_df = pd.read_excel(raw_file_path, dtype={"FAT_LONGITUDE" : str , "FAT_LATITUDE" : str, "BUILDING_LATITUDE" : str, "BUILDING_LONGITUDE" : str, "HOMEPASS_ID" : str, "RT" : str, "RW" : str}, sheet_name=hpdb_sheet[0])
    hpdb_df.columns = [col.replace(' ', '_') for col in hpdb_df.columns]

    # Add logic to update RFS_DATE and PARTNER_RFS_DATE columns
    try:
        for index, row in hpdb_df.iterrows():
            if row['RFS_DATE'] != "" and row['RFS_DATE'] != "-":
                hpdb_df.at[index, 'PARTNER_RFS_DATE'] = row['RFS_DATE']
            else:
                hpdb_df.at[index, 'RFS_DATE'] = row['PARTNER_RFS_DATE']    
    except:
        formatted_print(f'Error occurred while updating RFS_DATE and PARTNER_RFS_DATE columns', log_location, text_color)

    # Load City_Code.xlsx into a DataFrame
    city_code_df = pd.read_excel('Reference/City_Code.xlsx')

    #----------------------
    # City Code Lookup
    #----------------------
    # Merge the two DataFrames on the 'CITY' column
    try:
        merged_df = pd.merge(hpdb_df, city_code_df[['CITY', 'CITY_CODE']], on='CITY', how='left')
    except:
        formatted_print(f'Error occurred while merging DataFrames on CITY column', log_location, text_color)

    # Update CITY_CODE in HPDB SAMPEL.xlsx with the values from merged_df
    hpdb_df['CITY_CODE'] = merged_df['CITY_CODE_y']

    #----------------------
    # Mobile Region and Cluster Lookup
    #----------------------
    # Create a copy of the CITY column in hpdb_df
    try:
        hpdb_df['CITY_original'] = hpdb_df['CITY']

        # Convert CITY columns to lowercase for case-insensitive comparison
        hpdb_df['CITY'] = hpdb_df['CITY'].str.lower()
        mobile_df['CITY'] = mobile_df['CITY'].str.lower()

        # Merge the DataFrames on 'CITY' column
        merged_df = pd.merge(hpdb_df, mobile_df, on='CITY', how='left', suffixes=('', '_MOBILE'))

        # Replace the values in MOBILE_REGION and MOBILE_CLUSTER with the corresponding values from REGION and CLUSTER
        merged_df['MOBILE_REGION'] = merged_df['REGION MOBILE'].combine_first(merged_df['MOBILE_REGION'])
        merged_df['MOBILE_CLUSTER'] = merged_df['CLUSTER MOBILE'].combine_first(merged_df['MOBILE_CLUSTER'])

        # Drop the extra 'REGION' and 'CLUSTER' columns
        merged_df.drop(['REGION MOBILE', 'CLUSTER MOBILE', 'PROVINCE', 'CITY.1'], axis=1, inplace=True)

        # Replace CITY column with original values
        merged_df['CITY'] = merged_df['CITY_original']

        # Drop the 'CITY_original' column
        merged_df.drop('CITY_original', axis=1, inplace=True)
    except:
        formatted_print(f'Error occurred while processing Mobile Region and Cluster Lookup', log_location, text_color)

    #----------------------
    # Null Filling
    #----------------------
    try:
        # Replace null values with '-'
        hpdb_filled_df = merged_df.fillna('-')

        # Remove double spaces from all columns
        hpdb_filled_df = hpdb_filled_df.replace(r'\s+', ' ', regex=True)

        # Remove '|' characters from PROJECT_NAME column
        hpdb_filled_df['PROJECT_NAME'] = hpdb_filled_df['PROJECT_NAME'].str.replace('|', '')

        # Create new column based on concatenated columns
        hpdb_filled_df['ADDRESS'] = hpdb_filled_df[['PREFIX_ADDRESS', 'STREET_NAME', 'HOUSE_NUMBER', 'BLOCK', 'FLOOR', 'RT', 'RW']].apply(lambda x: ' '.join(x.astype(str)), axis=1)

        # Save the filled DataFrame to a new Excel file
        hpdb_filled_df.to_excel('Temp/temp.xlsx', index=False)

        # Load HPDB SAMPEL.xlsx into a DataFrame
        hpdb_df = hpdb_filled_df

        wb = load_workbook('Temp/temp.xlsx')
        ws = wb.active
    except:
        formatted_print(f'Error occurred while filling null values', log_location, text_color)

    #----------------------
    # RFS_DATE and PARTNER_RFS_DATE
    #----------------------
    # Find the column index of the date column
    try:
        date_column_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "RFS_DATE":  # Change "RFS_DATE" to the name of your date column
                date_column_index = idx
                break

        if date_column_index is not None:
            # Set the style of the column to a named style that displays only the date without time
            for row in ws.iter_rows(min_row=2, min_col=date_column_index, max_col=date_column_index):
                for cell in row:
                    # Mengubah objek datetime menjadi string dengan format baru 'dd/mm/yyyy'
                    if cell.value != "-":
                        cell.value = date_converter(cell.value)

        # Find the column index of the date column
        date_column_index = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "PARTNER_RFS_DATE":  # Change "RFS_DATE" to the name of your date column
                date_column_index = idx
                break

        if date_column_index is not None:
            # Set the style of the column to a named style that displays only the date without time
            for row in ws.iter_rows(min_row=2, min_col=date_column_index, max_col=date_column_index):
                for cell in row:
                    # Mengubah objek datetime menjadi string dengan format baru 'dd/mm/yyyy'
                    if cell.value != "-":
                        cell.value = date_converter(cell.value)
    except:
        formatted_print(f'Error occurred while processing RFS_DATE and PARTNER_RFS_DATE columns', log_location, text_color)

    #----------------------
    # List and Variables
    #----------------------
    # Initialize a PatternFill object for the red color
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Define the lists of allowed values
    acquisition_class_values = ["HOME", "HOME - BIZ", "BIZ - HOME", "BIZ"]
    building_type_values = ['PERUMAHAN', 'RUKO', 'FASUM']
    ownership_values = ['PARTNERSHIP-LN', 'PARTNERSHIP-IF', 'PARTNERSHIP-TBG', 'OWN BUILT']
    vendor_name_values = ['LINKNET', 'IFORTE', 'TBG']
    prefix_address_values = ['JL.', 'GG.']

    # Define the lists of similar columns
    coordinate_col = ['FDT_LONGITUDE', 'FAT_LONGITUDE', 'BUILDING_LONGITUDE', 'FDT_LATITUDE', 'FAT_LATITUDE', 'BUILDING_LATITUDE']

    # Convert the allowed values to lowercase
    acquisition_class_values_lower = [value.lower() for value in acquisition_class_values]
    building_type_values_lower = [value.lower() for value in building_type_values]
    ownership_values_lower = [value.lower() for value in ownership_values]
    vendor_name_values_lower = [value.lower() for value in vendor_name_values]
    prefix_address_values_lower = [value.lower() for value in prefix_address_values]

    # List to store row indices
    rows_to_delete = []

    #----------------------
    # Data Converting
    #----------------------
    # Convert string columns to lowercase
    try:
        hpdb_df = hpdb_df.map(lambda x: x.lower() if isinstance(x, str) else x)
        kodepos_df = kodepos_df.map(lambda x: str(x).lower() if isinstance(x, str) else x)
    except:
        formatted_print(f'Error occurred while converting string columns to lowercase', log_location, text_color)

    #-------------------------
    # LONGITUDE and LATITUDE
    #-------------------------
    # Replace all commas with dots and truncate decimals to 6 digits
    try:
        for col in coordinate_col:
            for row in ws.iter_rows(min_row=2, max_row=len(hpdb_df) + 1, min_col=hpdb_df.columns.get_loc(col) + 1, max_col=hpdb_df.columns.get_loc(col) + 1):
                for cell in row:
                    value = str(cell.value).replace(',', '.')  # Replace commas with dots
                    if '.' in value:
                        integer_part, decimal_part = value.split('.')
                        decimal_part = decimal_part.ljust(6, '0')[:6]  # Ensure decimal part has exactly 6 digits
                        value = f"{integer_part}.{decimal_part}"
                    cell.value = value
        
        homepass_long_lat["Coordinates"] = [[round(coord[0], 6), round(coord[1], 6)] for coord in homepass_long_lat["Coordinates"]]
        fat_long_lat["Coordinates"] = [[round(coord[0], 6), round(coord[1], 6)] for coord in fat_long_lat["Coordinates"]]

        for index, row in hpdb_df.iterrows():
            build_longitude = row['BUILDING_LONGITUDE']
            build_latitude = row['BUILDING_LATITUDE']
            fat_longitude = row['FAT_LONGITUDE']
            fat_latitude = row['FAT_LATITUDE']

            if pd.notnull(build_longitude) and pd.notnull(build_latitude) and has_hp:
                # Ubah nilai longitude dan latitude menjadi tuple
                coordinate_tuple = [float(build_longitude), float(build_latitude)]
                if not str(coordinate_tuple) in homepass_long_lat["Coordinates"].astype(str).to_list():
                    # Tambahkan logika untuk menandai jika nilai tidak ditemukan
                    ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LONGITUDE') + 1).fill = red_fill
                    ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LATITUDE') + 1).fill = red_fill

            elif pd.notnull(fat_longitude) and pd.notnull(fat_latitude) and has_fat:
                # Ubah nilai longitude dan latitude menjadi tuple
                coordinate_tuple = [float(fat_longitude), float(fat_latitude)]
                if not str(coordinate_tuple) in fat_long_lat["Coordinates"].astype(str).to_list():
                    # Tambahkan logika untuk menandai jika nilai tidak ditemukan
                    ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LONGITUDE') + 1).fill = red_fill
                    ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LATITUDE') + 1).fill = red_fill

    except Exception as e:
        formatted_print(f'Error occurred while processing LONGITUDE and LATITUDE columns. Error: {e}', log_location, text_color)

        for index, row in hpdb_df.iterrows():
            build_longitude = row['BUILDING_LONGITUDE']
            build_latitude = row['BUILDING_LATITUDE']
            fat_longitude = row['FAT_LONGITUDE']
            fat_latitude = row['FAT_LATITUDE']

            if not has_fat:
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LONGITUDE') + 1).fill = red_fill
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('FAT_LATITUDE') + 1).fill = red_fill
            
            elif not has_hp:
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LONGITUDE') + 1).fill = red_fill
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('BUILDING_LATITUDE') + 1).fill = red_fill

    #----------------------
    # Zip Code Checking
    #----------------------
    try:
        # Iterate over each row in the "HPDB SAMPEL" DataFrame
        for index, row in hpdb_df.iterrows():
            # Check if there is a matching row in the "kodepos" DataFrame
            match = kodepos_df[
                (kodepos_df['REGION'] == row['REGION']) &
                (kodepos_df['CITY'] == row['CITY']) &
                (kodepos_df['DISTRICT'] == row['DISTRICT']) &
                (kodepos_df['SUB_DISTRICT'] == row['SUB_DISTRICT']) &
                (kodepos_df['ZIP_CODE'].astype(str) == str(row['ZIP_CODE']))
            ]
            # If a match is found, mark the row in the Excel worksheet for coloring
            if match.empty:
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('REGION') + 1, max_col=hpdb_df.columns.get_loc('ZIP_CODE') + 1):
                    for c in cell:
                        c.fill = red_fill
    except:
        formatted_print(f'Error occurred while processing ZIP_CODE column', log_location, text_color)

    #----------------------
    # ACQUISITION_CLASS
    #----------------------
    try:
        # Fill all existing rows with red color if ACQUISITION_CLASS is not in allowed_values
        for index, row in hpdb_df.iterrows():
            if row['ACQUISITION_CLASS'].lower() not in acquisition_class_values_lower:
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1):
                    for c in cell:
                        c.fill = red_fill
    except:
        formatted_print(f'Error occurred while processing ACQUISITION_CLASS column', log_location, text_color)

    #----------------------
    # BUILDING_TYPE
    #----------------------
    try:
        # Iterate over each row in the "HPDB SAMPEL" DataFrame
        for index, row in hpdb_df.iterrows():
            # Check if BUILDING_TYPE is one of the allowed values
            if row['BUILDING_TYPE'] not in building_type_values_lower:
                # Mark the row in the Excel worksheet for coloring
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1, max_col=hpdb_df.columns.get_loc('BUILDING_TYPE') + 1):
                    for c in cell:
                        c.fill = red_fill
    except:
        formatted_print(f'Error occurred while processing BUILDING_TYPE column', log_location, text_color)

    #----------------------
    # ACQUISITION_TIER
    #----------------------
    try:
        # Iterate over each row in the "HPDB SAMPEL" DataFrame
        for index, row in hpdb_df.iterrows():
            # Check if BUILDING_TYPE is one of the allowed values
            if row['BUILDING_TYPE'] == "ruko":
                #
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1):
                    for c in cell:
                        c.value = "HOME - BIZ"

                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_CLASS') + 1):
                    for c in cell:
                        c.value = "BIZ"
            else:
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1, max_col=hpdb_df.columns.get_loc('ACQUISITION_TIER') + 1):
                    for c in cell:
                        c.value = "HOME"
    except:
        formatted_print(f'Error occurred while processing ACQUISITION_TIER column', log_location, text_color)

    #----------------------
    # OWNERSHIP
    #----------------------
    try:
        # Check if any of the ownership values is in the OWNERSHIP field
        for index, row in hpdb_df.iterrows():
            found = False
            for value in ownership_values_lower:
                if value in row['OWNERSHIP']:
                    found = True
                    break
            if not found:
                # Fill the cell with red color
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1, max_col=hpdb_df.columns.get_loc('OWNERSHIP') + 1):
                    for c in cell:
                        c.fill = red_fill
            elif row['OWNERSHIP'] != 'OWN BUILT' and row['VENDOR_NAME'] not in vendor_name_values_lower:
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1, max_col=hpdb_df.columns.get_loc('VENDOR_NAME') + 1):
                    for c in cell:
                        c.fill = red_fill
    except:
        formatted_print(f'Error occurred while processing OWNERSHIP column', log_location, text_color)

    #----------------------
    # PREFIX_ADDRESS
    #----------------------
    try:
        for index, row in hpdb_df.iterrows():
            if row['PREFIX_ADDRESS'] not in prefix_address_values_lower:
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1):
                    for c in cell:
                        c.fill = red_fill
    except:
        formatted_print(f'Error occurred while processing PREFIX_ADDRESS column', log_location, text_color)

    #--------------------------------------------------------
    # CLUSTER_NAME, CLUSTER_CODE, PROJECT_NAME, CITY_CODE
    #--------------------------------------------------------
    # Iterate over each row in the DataFrame
    for index, row in hpdb_df.iterrows():
        try:
            # Check if the CLUSTER_NAME column exceeds 100 characters
            if len(str(row['CLUSTER_NAME'])) > 100 or row['CLUSTER_NAME'] == "-":
                # Mark the row in the Excel worksheet for coloring
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_NAME') + 1):
                    for c in cell:
                        c.fill = red_fill

            # Check if the CLUSTER_CODE column exceeds 20 characters
            if len(str(row['CLUSTER_CODE'])) > 20 or row['CLUSTER_CODE'] == "-":
                # Mark the row in the Excel worksheet for coloring
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1, max_col=hpdb_df.columns.get_loc('CLUSTER_CODE') + 1):
                    for c in cell:
                        c.fill = red_fill

            # Check if the length of PROJECT_NAME exceeds 100 characters
            if len(row['PROJECT_NAME']) > 100 or row['PROJECT_NAME'] == "-":
                # Fill the entire row with red color
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1, max_col=hpdb_df.columns.get_loc('PROJECT_NAME') + 1):
                    for c in cell:
                        c.fill = red_fill

            # Check if the length of PROJECT_NAME exceeds 100 characters
            if row['CITY_CODE'] == "-":
                # Fill the entire row with red color
                for cell in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('CITY_CODE') + 1, max_col=hpdb_df.columns.get_loc('CITY_CODE') + 1):
                    for c in cell:
                        c.fill = red_fill
        except:
            formatted_print(f'Error occurred while processing CLUSTER_NAME, CLUSTER_CODE, PROJECT_NAME, CITY_CODE columns', log_location, text_color)

    #-------------
    # HOMEPASS_ID
    #-------------
    try:
        # Get the duplicated rows in the "HOMEPASS_ID" column
        duplicates = hpdb_df["HOMEPASS_ID"].duplicated()

        # Fill the duplicated rows with red color
        for index, duplicate in duplicates.items():
            if duplicate:
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('HOMEPASS_ID') + 1).fill = red_fill

        # Iterate over each row in the DataFrame
        for index, row in hpdb_df.iterrows():
            # Check if the PARTNER_RFS_DATE column is NULL or has an incorrect format
            if pd.isnull(row['PARTNER_RFS_DATE']) or \
                not isinstance(row['PARTNER_RFS_DATE'], pd.Timestamp) or \
                row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y') != row['PARTNER_RFS_DATE'].strftime('%m/%d/%Y'):
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('PARTNER_RFS_DATE') + 1).fill = red_fill

            # Check if the RFS_DATE column is NULL or has an incorrect format
            if pd.isnull(row['RFS_DATE']) or \
                not isinstance(row['RFS_DATE'], pd.Timestamp) or \
                row['RFS_DATE'].strftime('%m/%d/%Y') != row['RFS_DATE'].strftime('%m/%d/%Y'):
                ws.cell(row=index + 2, column=hpdb_df.columns.get_loc('RFS_DATE') + 1).fill = red_fill
    except:
        formatted_print(f'Error occurred while processing HOMEPASS_ID column', log_location, text_color)

    #-----------------------------
    # Address Duplicate Checking
    #-----------------------------
    try:
        duplicates = hpdb_filled_df["ADDRESS"].duplicated()

        # Fill the duplicated rows with red color
        for index, duplicate in duplicates.items():
            if duplicate:
                for row in ws.iter_rows(min_row=index + 2, max_row=index + 2, min_col=hpdb_df.columns.get_loc('PREFIX_ADDRESS') + 1, max_col=hpdb_df.columns.get_loc('RW') + 1):
                    for cell in row:
                        cell.fill = red_fill

        # Drop the "ADDRESS" column
        ws.delete_cols(hpdb_filled_df.columns.get_loc("ADDRESS") + 1, amount=1)
    except:
        formatted_print(f'Error occurred while processing ADDRESS column', log_location, text_color)

    #-------------
    # Save
    #-------------
    # Save the modified Excel workbook
    wb.save(output_file_path)
    formatted_print(f'Output file is saved', log_location, text_color)

    #-----------------------------
    # Summary
    #-----------------------------
    try:
        # Load the Excel workbook
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Iterate over each row starting from the second row
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # Check if any cell in the row has red fill
            has_red_fill = any([cell.fill == red_fill for cell in row])

            # If no red fill, add row index to rows_to_delete list
            if not has_red_fill:
                rows_to_delete.append(row[0].row)

        # Delete rows in reverse order to avoid shifting rows
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

        # Iterate over each cell in the worksheet
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Skip certain columns
                if ws.cell(row=1, column=cell.column).value in ['HOMEPASS_ID']:
                    continue
                
                if cell.fill == red_fill:
                    cell.value = 'REVISE'

                else:
                    cell.value = 'OK'

        wb.save('Temp/HPDB Summary.xlsx')

        # Create a DataFrame with column names from column_names
        hpdb_summary_df = pd.read_excel('Temp/HPDB Summary.xlsx')
        kmz_summary_df = pd.DataFrame(columns=kmz_col)

        # Menghapus kolom yang tidak ada dalam list
        columns_to_drop = set(hpdb_summary_df.columns) - set(hpdb_col)
        hpdb_summary_df.drop(columns=columns_to_drop, inplace=True)

        hpdb_summary_num_rows = hpdb_summary_df.shape[0]

        # Create a new row with "-" values
        new_row = pd.DataFrame([["-"] * len(kmz_summary_df.columns)], columns=kmz_summary_df.columns)

        # Append the new row to kmz_summary_df num_rows times
        for _ in range(hpdb_summary_num_rows):
            kmz_summary_df = pd.concat([kmz_summary_df, new_row], ignore_index=True)

        # Assuming log_col, cluster, checking_date, and checking_time are defined elsewhere in your code
        log_summary_df = pd.DataFrame(columns=log_col)
        log_summary_df_val = [cluster, checking_date, checking_time, "REVISE"]

        # Create a DataFrame from log_summary_df_val and repeat it for hpdb_summary_num_rows times
        new_log_rows = pd.DataFrame([log_summary_df_val] * (hpdb_summary_num_rows), columns=log_col)

        # Append the new log rows to log_summary_df
        log_summary_df = pd.concat([log_summary_df, new_log_rows], ignore_index=True)

        summary_df = pd.concat([log_summary_df, kmz_summary_df, hpdb_summary_df], axis=1, ignore_index=False)
        summary_df.to_csv(summary_file_path_cluster, index=False) #summary per cluster (only last for 5 days)

        if not os.path.exists(summary_file_path):
            # Save DataFrame to CSV if the file does not exist
            summary_df.to_csv(summary_file_path, index=False)
            formatted_print(f'Summary log CSV file is created', log_location, text_color)
        else:
            # Append DataFrame to existing CSV file
            summary_df.to_csv(summary_file_path, mode='a', header=False, index=False)
            formatted_print(f'Summary log is appended to the CSV file', log_location, text_color)
    except Exception as e :
        formatted_print(f'Error occurred while processing Summary. Error: {e}', log_location, text_color)

    #-------------
    # Finishing
    #-------------
    end_time = time.time()
    execution_time = end_time - start_time
    formatted_print(f'Process was done in {execution_time:.2f} seconds', log_location, text_color)

def main():
    hpdb_file = "HPDB-XL35GSK01160_001.xlsx"
    kmz_file = "ABD-XL35GSK01160_001.kmz"
    cluster_id = "XL35GSK01160_001"
    log_col = ['Cluster ID', 'Checking Date', 'Checking Time', "Status"]
    kmz_col = ['Pole to FAT', 'Pole to FDT', 'HP to all infrastructure 35m', 'Coordinate HP to all infrastructure 35m', 'HP to FAT 150m', 'Coordinate HP to FAT 150m',
            "Pole not in Distribution and Sling", "Coordinate Pole not in Distribution and Sling"]
    hpdbCheck(f'Input\{cluster_id}\{hpdb_file}', f'Input\{cluster_id}\{kmz_file}', f'{cluster_id}', "2022-01-01", "00:00:00", kmz_col, log_col)

if __name__ == '__main__':
    main()
